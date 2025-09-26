import pandas as pd
import io
from typing import Dict, Any, List
from openpyxl.utils import get_column_letter
import calendar
from docx import Document

import pandas as pd
import io
from typing import Dict, Any, List
from openpyxl.utils import get_column_letter
import calendar

def read_input_file(file_data: bytes, file_extension: str) -> pd.DataFrame:
    """Read and parse input file (CSV or Excel) while preserving original order."""
    try:
        if file_extension.lower() == '.csv':
            df = pd.read_csv(io.BytesIO(file_data), encoding='utf-8')
        else:  # Excel
            df = pd.read_excel(io.BytesIO(file_data))

        # Add original order index
        df['original_order'] = range(len(df))

        # Clean up the dataframe while preserving order
        df = df[['Title', 'Permalink', 'Durata Curs', 'original_order']].dropna(subset=['Title', 'Durata Curs'])
        
        # Extract numeric duration
        df['duration'] = df['Durata Curs'].str.extract('(\d+)').astype(int)

        return df
    except Exception as e:
        raise ValueError(f"Error reading file: {str(e)}")

def create_excel_export(schedule: List[Dict[str, Any]], year: int, holidays: List[str] = None) -> bytes:
    """Create Excel file from schedule data with original order and holidays tab."""
    try:
        output = io.BytesIO()
        
        # Create DataFrame first using a list to maintain order
        course_list = []
        seen_titles = set()  # To track which courses we've seen
        
        # Sort once by original order
        sorted_schedule = sorted(schedule, key=lambda x: x.get('original_order', 0))
        
        for item in sorted_schedule:
            title = item['Title']
            if title not in seen_titles:
                seen_titles.add(title)
                course_list.append({
                    'Title': title,
                    'Permalink': item['Permalink'],
                    'Durata Curs': item['Durata Curs'],
                    'January': '',
                    'February': '',
                    'March': '',
                    'April': '',
                    'May': '',
                    'June': '',
                    'July': '',
                    'August': '',
                    'September': '',
                    'October': '',
                    'November': '',
                    'December': ''
                })
            
            # Find the course in our list and update its month
            course_entry = next(c for c in course_list if c['Title'] == title)
            month_name = calendar.month_name[item['month']]
            course_entry[month_name] = item['date_range']

        # Convert list to DataFrame - order will be preserved
        schedule_df = pd.DataFrame(course_list)
        
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            schedule_df.to_excel(writer, sheet_name='Schedule', index=False)
            
            if holidays:
                holidays_df = pd.DataFrame(holidays, columns=['Holiday Date'])
                holidays_df.to_excel(writer, sheet_name='Holidays', index=False)

            for sheet in writer.sheets.values():
                for idx, col in enumerate(sheet.iter_cols()):
                    max_length = max(len(str(cell.value)) for cell in col)
                    sheet.column_dimensions[get_column_letter(idx + 1)].width = max_length + 2

        output.seek(0)
        return output.getvalue()
    
    except Exception as e:
        raise ValueError(f"Error creating Excel file: {str(e)}")
    


from docx import Document
import pandas as pd
import io

def process_word_file(file_data: bytes) -> dict:
    """Extract Luna columns from Word document and return a mapping of course titles to their date ranges."""
    document = Document(io.BytesIO(file_data))
    course_dates = {}
    
    for table in document.tables:
        for row in table.rows:
            cells = row.cells
            
            # Skip rows with merged headers spanning multiple columns
            merged_check = {cell._tc for cell in cells}
            if len(merged_check) == 1:
                continue
            
            # Extract cell text and ensure there are enough columns
            cell_texts = [cell.text.strip() for cell in cells]
            if len(cell_texts) >= 6:  # Ensure we have enough columns including price
                course_title = cell_texts[0]
                # Skip duration (cell_texts[1]) and price (cell_texts[2])
                luna_1 = cell_texts[3]  # Luna 1 is now at index 3
                luna_2 = cell_texts[4]  # Luna 2 is now at index 4
                luna_3 = cell_texts[5]  # Luna 3 is now at index 5
                
                # Store only if we have a title and at least one luna value
                if course_title and any([luna_1, luna_2, luna_3]):
                    course_dates[course_title] = {
                        'Luna 1': luna_1,
                        'Luna 2': luna_2,
                        'Luna 3': luna_3
                    }
    
    print(f"Extracted date ranges for {len(course_dates)} courses from Word document")
    return course_dates

def convert_to_excel(df: pd.DataFrame) -> bytes:
    """Convert DataFrame to Excel format."""
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Courses')
    output.seek(0)
    return output.getvalue()


def convert_word_to_excel(file_data: bytes) -> bytes:
    """Convert the processed Word document data to Excel format."""
    df = process_word_file(file_data)
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Courses')
    output.seek(0)
    return output.getvalue()
